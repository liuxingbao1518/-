from openpyxl import load_workbook

# 安装命令
# pip install openpyxl
wb=load_workbook('./test.xlsx')
# 输出整个excel表格的表单名
# print(wb.sheetnames)
ws=wb['Sheet2']
# print(ws.cell(row=5,column=1).value)

# 获取表单中有数据的总行数
# print(ws.max_row)
# 获取表单中有数据的总列数
# print(ws.max_column)
# 修改数据
ws.cell(row=5,column=1).value='老婆是余丽华005'
# 注意哪个对象打开的，就哪个对象保存
wb.save('./test.xlsx')
print(ws.cell(row=5,column=1).value)
wb.close()

