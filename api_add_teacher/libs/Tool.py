import os
import time
import yagmail
import pymysql
from selenium import webdriver
from openpyxl import load_workbook
from config.Secert import *
import logging
import requests
import unittest
# 发送邮件
def send_email(user,password,port,body,subject,report,to_user,host='smtp.163.com'):
    """

    :param user:发送者
    :param password:授权码
    :param port:端口默认为3306
    :param body:主题
    :param subject:标题
    :param report:报告
    :param to_user:接受者账号，默认是字符串，如果传多个请用列表的方式传递
    :param host:邮箱服务
    :return:
    """
    # 生成发送对象
    send = yagmail.SMTP(user=user, password=password, host=host, port=port)
    if type(to_user) is list:
        send.send(to=to_user, cc=user, subject=subject, contents=[body, report])
        flag = '发送给批量用户成功'
    elif type(to_user) is str:
        send.send(to=to_user, cc=user, subject=subject, contents=[body, report])
        flag = '发送个人用户成功'

    else:
        flag = '发送数据错误'
    return flag

FD = "./reports"
# 获取最新的测试报告
def GetNewReport(FileDir=FD):
    # 打印目录所在所有文件名（列表对象）
    l = os.listdir(FileDir)
    # 按时间排序
    l.sort(key=lambda fn: os.path.getmtime(FileDir + "\\" + fn))
    # 获取最新的文件保存到file_new
    f = os.path.join(FileDir, l[-1])
    return f




# 数据库的操作
def read_mysql_data(host,port,user,password,db,sql):
    """
    :param host:本地服务
    :param port:
    :param user:
    :param password:
    :param db: 数据库名称
    :param sql:
    :return:
    """
    # 建立链接
    conn = pymysql.Connect(host=host, port=port, user=user, password=password, db=db, )
    # 生成游标
    cur = conn.cursor()
    # 执行sql语句
    cur.execute(sql)
    # 关闭游标
    cur.close()
    # 关闭数据库
    conn.close()
    # 查询sql结果
    d = cur.fetchone()
    return d

# 获取测试数据
def getTestData(local, requirement, casename):
    """
    :param local: 加载测试用例的路径
    :param requirement:测试需求（对应表单）
    :param casename:测试用例（对应编号）
    :return:
    """
    # 加载excel表格
    wb = load_workbook(local)
    # 加载对应的测试需求（表单）
    ws = wb[requirement]
    for i in range(len(ws['A'])):
        # 判断的A列下面的编号名字如果==传入来的测试用例的名字就取对应的值
        if ws['A'][i].value == casename:
            # 取出需要的第五行测试数据的数据
            result = ws.cell(row=ws['A'][i].row,column=5).value
            if result is not None:
                # 拿到数据使用','来把数据分割成列表
                stuff = result.find(',')
                # 查找数据中有没有，如果没有就会返回-1
                if stuff != -1:
                    data = result.split(',')
                    # 返回的是一个列表
                    return data
    wb.close()

#创建日志
# 当前文件路径
cur_path = os.path.dirname(os.path.realpath(__file__))

# log_path是存放日志的路径
log_path = os.path.join(os.path.dirname(cur_path), 'log')
# 如果不存在这个log文件夹，就自动创建一个
if not os.path.exists(log_path):
    os.mkdir(log_path)


class InsertLog(object):
    def __init__(self):
        # 文件的命名
        self.logname = os.path.join(log_path, '%s.log' % time.strftime('%Y_%m_%d'))
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        # 日志输出格式
        self.formatter = logging.Formatter(
            '[%(asctime)s - %(funcName)s line: %(lineno)3d] - %(levelname)s: %(message)s')

    def __console(self, level, message,*args):
        # 创建一个FileHandler，用于写到本地
        #fh = logging.FileHandler(self.logname, 'a')  # 追加模式  这个是python2的
        fh = logging.FileHandler(self.logname, 'a', encoding='utf-8')  # 这个是python3的
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(self.formatter)
        self.logger.addHandler(fh)

        # 创建一个StreamHandler,用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(self.formatter)
        self.logger.addHandler(ch)

        if level == 'info':
            self.logger.info(message)
        elif level == 'debug':
            self.logger.debug(message)
        elif level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)
        # 这两行代码是为了避免日志输出重复问题
        self.logger.removeHandler(ch)
        self.logger.removeHandler(fh)
        # 关闭打开的文件
        fh.close()

    def debug(self, message,*args):
        self.__console('debug', message,args)

    def info(self, message,*args):
        self.__console('info', message,args)

    def warning(self, message,*args):
        self.__console('warning', message,args)

    def error(self, message,*args):
        self.__console('error', message,args)




#请求接口返回结果
class BaseHttp(object):
    #定义类变量
    host = host

    def http_send(self,url,method='post',*args,**kwargs):
        url = '{}{}'.format(self.host, url)
        # print('url is',url)
        # print('method is',method)
        result = requests.request(method=method,url=url,*args,**kwargs)
        return result



#接口校验
class VerifyClass(unittest.TestCase):
    #校验html字段
    def verify_html(self,result,v_html):
        '''
        :param result: 请求的结果
        :param v_html: 校验的结果值
        :return:
        '''
        self.assertIn(v_html,result.text)

    #校验状态码
    def verify_code(self,result,v_code):
        self.assertEqual(result.status_code,v_code)

    #校验响应体（只有响应体是json的时候才可以用）
    def verify_json(self,result,v_json):
        self.assertEqual(result.json(),v_json)

    #综合校验多个字段
    def verify_union(self,result,v_html,v_code):
        self.verify_html(result,v_html)
        self.verify_code(result,v_code)

    #校验接口与接口之间的依赖
    #校验登陆接口与充值接口的依赖
    def verify_login_add_union(self,result,v_html,v_code=200):
        # 校验登录
        self.verify_code(result[0],v_code)

        # 校验充值
        self.verify_code(result[1],v_code)
        self.verify_html(result[1],v_html)



if __name__ == '__main__':
        pass